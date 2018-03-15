from openpyxl import Workbook, load_workbook
from openpyxl.worksheet import Worksheet
from datetime import datetime
from app.src import models
from typing import List


def json_to_clustering_result(json) -> models.InputClustering:
    id = json["id"]
    clusters = list(map(lambda e: json_to_cluster(e), json["clusters"]))
    return models.InputClustering(id, clusters)


def json_to_cluster(json) -> models.InputCluster:
    id = json["id"]
    categories = json["categories"]
    return models.InputCluster(id, categories)


def multiple_results_from_json(json) -> List[models.InputClustering]:
    """
    Takes in multiple json files which contain InputClustering-type data
    :param json:
    :return:
    """
    all_crs = []
    for res_obj in json:
        o = list(map(lambda obj: json_to_clustering_result(obj), res_obj))
        all_crs.append(o)
    return all_crs

def get_lengths(cl):
    ct = extract_category_titles(cl)
    cat_len = len(ct)
    clus_len = len(cl.clusters)
    return cat_len, clus_len

def extract_category_titles(cl):
    category_titles = set(cl.clusters[0].categories)
    return category_titles

def append_results_to_excel(cls: List[models.InputClustering], workbook) -> Workbook:
    """
    Appends results to workbook with current time as suffix
    :param cls: list of models.InputClustering
    :param workbook:
    :return:
    """
    sheet_title = f"Clusters {1+1}"  # {datetime.now() # could be parametrized
    wb = Workbook()
    ws: Worksheet = wb.create_sheet(sheet_title,0) # len(wb.worksheets())

    cat_len = 0
    clus_len = 0

    c_title_start_row = 3
    c_title_column = 1
    start_row = 4
    clusters_start_col = 2
    padding = 1

    category_titles = set()

    i = 0
    row_cursor = start_row
    for cl in cls:
        cluster_id = cl.id


        # set the basic lengths
        if cat_len == 0 and clus_len == 0:
            (cat_len, clus_len) = get_lengths(cl)

        # get the category titles
        if len(category_titles) == 0:
            category_titles = extract_category_titles(cl)

        first_pass = True
        for cluster in cl.clusters:

            # write titles for columns
            k = 0
            if first_pass:
                ws.cell(row=row_cursor, column=c_title_column, value=cluster_id)
                for cat_title in category_titles:
                    ws.cell(row=row_cursor, column=clusters_start_col + k, value=cat_title)
                    k += 1
                row_cursor += 1


            k = 0
            for cat in category_titles:
                cat_count = cluster.categories[cat]
                ws.cell(row=row_cursor, column=clusters_start_col + k, value=cat_count)
                k += 1
            row_cursor += 1
            first_pass = False
        row_cursor += padding
        i += 1


    wb.save('document.xlsx')
