from typing import List

from openpyxl import Workbook
from openpyxl.worksheet import Worksheet

import os

from app.src import excel_formatting_utils as formatting
from app.src.excel_formatting_utils import Direction
from app.src import models


def extract_category_titles_from_cluster(cl):
    category_titles = sorted(set(cl.clusters[0].categories))
    return category_titles


class FormattingData(object):
    start_column = 0
    start_row = 0
    padding = 0

    def __init__(self, start_column,
                 start_row,
                 padding):
        self.start_column = start_column
        self.start_row = start_row
        self.padding = padding


def multiple_results_to_excels(base_fd: FormattingData, results, filename=None):
    """

    :param base_fd: FormattingData
    :param results: List of clustering results from multiple files
    :return:
    """
    wb = Workbook()
    ws = wb.create_sheet("Result", 0)
    calculate_entropy_for_cluster(results)
    cluster_averages = get_cluster_averages(results)
    col_cursor = base_fd.start_column
    len_results = len(results)
    for i, result in enumerate(results):
        fd = FormattingData(col_cursor, base_fd.start_row, base_fd.padding)
        if i + 1 == len_results:
            (cc, rc) = add_results_to_excel(fd, result, ws, cluster_averages)
        else:
            (cc, rc) = add_results_to_excel(fd, result, ws)
        col_cursor = cc  # This just increases to column by 4 for each iteration

    par_dir = os.path.dirname(filename)
    if not os.path.exists(par_dir):
        os.makedirs(par_dir)
    if filename is not None:
        return wb.save(f"{filename}")
    return wb.save(f"results.xlsx")


def get_cluster_averages(results: List[List[models.InputClustering]]):
    cluster_averages = {}
    avg_sil_f, avg_org_f, avg_pur_f, avg_ent_f, avg_runt_f = \
        "avg. silhouette", "avg. org. silhouette (sample)", "avg. purity score", "avg. entropy", "avg. running time (ms)"
    for result in results:
        for cls in result:
            if cls.id not in cluster_averages:
                cluster_averages[cls.id] = {
                    avg_sil_f: cls.silhouette,
                    avg_org_f: cls.original_silhouette,
                    avg_pur_f: cls.purity_score,
                    avg_ent_f: cls.entropy,
                    avg_runt_f: cls.running_time_ms
                }
            else:
                cluster_averages[cls.id][avg_sil_f] = \
                    (cluster_averages[cls.id][avg_sil_f] + cls.silhouette) / 2
                cluster_averages[cls.id][avg_org_f] = \
                    (cluster_averages[cls.id][avg_org_f] + cls.original_silhouette) / 2
                cluster_averages[cls.id][avg_pur_f] = \
                    (cluster_averages[cls.id][avg_pur_f] + cls.purity_score) / 2
                cluster_averages[cls.id][avg_ent_f] = \
                    (cluster_averages[cls.id][avg_ent_f] + cls.entropy) / 2
                cluster_averages[cls.id][avg_runt_f] = \
                    (cluster_averages[cls.id][avg_runt_f] + cls.running_time_ms) / 2
    return cluster_averages


def calculate_entropy_for_cluster(results: List[List[models.InputClustering]]):
    from math import log2
    for result in results:
        for cls in result:
            entropies_and_weights = []
            total = 0
            for cl in cls.clusters:
                items = cl.categories
                weight = sum(items.values())
                es = 0.0
                for i in items.values():
                    if i != 0:
                        e = i / weight * log2(i / weight)
                        es += e
                entropies_and_weights.append((es, weight))
                total += weight
            relative_entropies = []
            for (entropy, weight) in entropies_and_weights:
                relative_entropies.append(entropy * (weight / total))
            cls.entropy = sum(relative_entropies)

def write_cluster_averages(averages, ws: Worksheet, start_row, target_col):
    row_cursor = start_row
    row_cursor += 1
    for (k, v) in averages.items():
        name_cell = ws.cell(row=row_cursor, column=target_col, value=k)
        formatting.set_column_width(ws, name_cell.column, 20)
        ws.cell(row=row_cursor, column=target_col + 1, value=round(v, 4))
        row_cursor += 1
    return row_cursor


def write_cluster_result_stats(fd: FormattingData, cl: models.InputClustering, ws: Worksheet, start_row):
    # Write stats
    # Reset column cursor
    col_cursor = fd.start_column
    row_cursor = start_row
    row_cursor += 1
    first_cell = ws.cell(row=row_cursor, column=col_cursor, value="silhouette")
    formatting.add_border_to_row(ws, first_cell.row, Direction.TOP, formatting.borders.BORDER_DASHED)
    ws.cell(row=row_cursor, column=col_cursor + 1, value=round(cl.silhouette, 4))

    row_cursor += 1
    ws.cell(row=row_cursor, column=col_cursor, value="org. silhouette (sample)")
    ws.cell(row=row_cursor, column=col_cursor + 1, value=round(cl.original_silhouette, 4))

    row_cursor += 1
    ws.cell(row=row_cursor, column=col_cursor, value="purity score")
    ws.cell(row=row_cursor, column=col_cursor + 1, value=round(cl.purity_score, 4))

    row_cursor += 1
    ws.cell(row=row_cursor, column=col_cursor, value="entropy")
    ws.cell(row=row_cursor, column=col_cursor + 1, value=round(cl.entropy, 4))

    row_cursor += 1
    ws.cell(row=row_cursor, column=col_cursor, value="running time (ms)")
    ws.cell(row=row_cursor, column=col_cursor + 1, value=round(cl.running_time_ms, 4))

    return row_cursor + 1


def write_cluster_labels(cl: models.InputClustering, ws: Worksheet, row_start, target_col):
    row_cursor = row_start
    for clus in cl.clusters:
        cell = ws.cell(row=row_cursor, column=target_col, value=f"{clus.id}")
        formatting.right_align(cell)
        row_cursor += 1
    return row_cursor


def add_results_to_excel(fd: FormattingData,
                         cls: List[models.InputClustering],
                         ws: Worksheet,
                         cluster_averages=None):
    """
    Appends results to workbook with current time as suffix
    :param fd:
    :param cls: list of models.InputClustering
    :param ws:
    :return:
    """

    category_titles = set()

    row_cursor = fd.start_row
    col_cursor = 0
    title_column = None
    category_columns = []
    for cl in cls:
        col_cursor = fd.start_column
        cluster_id = cl.id

        # get the category titles
        if len(category_titles) == 0:
            category_titles = extract_category_titles_from_cluster(cl)

        header_pass = True
        for cluster in cl.clusters:
            k = 0
            if header_pass:
                # Write cluster titlte
                title_cell = ws.cell(row=row_cursor, column=col_cursor, value=cluster_id)
                formatting.bold_cell(title_cell)
                if title_column is None:
                    title_column = title_cell.column

                col_cursor += 1
                # Write categories
                for cat_title in category_titles:
                    cat_cell = ws.cell(row=row_cursor, column=col_cursor + k, value=cat_title)
                    formatting.bold_cell(cat_cell)
                    if len(category_columns) <= len(category_titles):
                        category_columns.append(cat_cell.column)
                    k += 1
                formatting.add_thin_border_to_row(ws, row_cursor, Direction.BOTTOM)
                row_cursor += 1

                # Write cluster labels
                write_cluster_labels(cl, ws, row_cursor, fd.start_column)

            k = 0
            # Write results per category
            for cat in category_titles:
                cat_count = cluster.categories[cat]
                ws.cell(row=row_cursor, column=col_cursor + k, value=cat_count)
                k += 1

            row_cursor += 1
            header_pass = False

        # Write stats
        stat_cursor = row_cursor # Just to store this for averages
        row_cursor = write_cluster_result_stats(fd, cl, ws, row_cursor)
        if cluster_averages is not None:
            write_cluster_averages(cluster_averages[cl.id], ws, stat_cursor, fd.start_column + len(category_titles) + 1)
            print(cluster_averages[cl.id])
        row_cursor += fd.padding

    col_cursor += len(category_titles)

    # Set some basic formatting
    formatting.set_column_width(ws, title_column, 20)
    formatting.add_thin_border_to_column(ws, title_column, Direction.RIGHT)
    for c in category_columns:
        formatting.set_column_width(ws, c, 15)
        if c == c[-1:]:
            formatting.add_thin_border_to_column(ws, title_column, Direction.RIGHT)

    return col_cursor, row_cursor
