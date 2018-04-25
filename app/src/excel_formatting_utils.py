from openpyxl.utils import column_index_from_string
from enum import Enum

from openpyxl.styles import Font, Alignment
from openpyxl.styles import borders
from openpyxl.utils import column_index_from_string
from copy import copy


def bold_cell(cell):
    f: Font = copy(cell.font)
    f.bold = True
    cell.font = f


def set_column_width(ws, col, width):
    ws.column_dimensions[col].width = width


class Direction(Enum):
    LEFT = 1
    TOP = 2
    RIGHT = 3
    BOTTOM = 4


def add_border(cell, direction, side):
    o = copy(cell.border)
    if direction == Direction.LEFT:
        o.left = side
    if direction == Direction.TOP:
        o.top = side
    if direction == Direction.RIGHT:
        o.right = side
    if direction == Direction.BOTTOM:
        o.bottom = side
    cell.border = o


def add_thin_border(cell, direction):
    # TODO Remove and use add_border_to_row
    side: borders.Side = borders.Side(border_style=borders.BORDER_THIN, color='000000')
    add_border(cell, direction, side)


def add_border_to_row(ws, row_index, direction, border):
    side: borders.Side = borders.Side(border_style=border, color='000000')
    for row in ws.iter_rows(min_row=row_index, max_row=row_index):
        for cell in row:
            add_border(cell, direction, side)


def add_thin_border_to_row(ws, row_index, direction):
    # TODO Remove and use add_border_to_row
    for row in ws.iter_rows(min_row=row_index, max_row=row_index):
        for cell in row:
            add_thin_border(cell, direction)


def add_thin_border_to_column(ws, column_letter, side):
    s: borders.Side = borders.Side(border_style=borders.BORDER_THIN, color='000000')

    b: borders.Border = None
    idx = column_index_from_string(column_letter)

    for column in ws.iter_cols(idx):
        for cell in column:
            add_thin_border(cell, side)


def right_align(cell):
    cell.alignment = Alignment(horizontal='right')
